#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""2024年研究生数学建模竞赛C题第一问：励磁波形分类。

本程序一次完成：
1. 读取附件一的四种材料，并按位置统一1024个磁通密度采样点；
2. 对每条波形进行去直流和峰峰值归一化；
3. 提取可解释的分布、形状和频域特征；
4. 进行分层五折验证和留一材料验证；
5. 使用全部训练数据预测附件二；
6. 将结果写入附件四第2列，同时输出验证报告和图表。
"""

from __future__ import annotations

import argparse
import json
import os
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

# 为只读主目录、CI和无界面环境提供可写缓存目录。
_runtime_cache = Path(tempfile.gettempdir()) / "c2024_waveform_cache"
_runtime_cache.mkdir(parents=True, exist_ok=True)
os.environ.setdefault("MPLCONFIGDIR", str(_runtime_cache / "matplotlib"))
os.environ.setdefault("XDG_CACHE_HOME", str(_runtime_cache))

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.stats import kurtosis, skew
from sklearn.base import clone
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    accuracy_score,
    classification_report,
    confusion_matrix,
    f1_score,
)
from sklearn.model_selection import (
    LeaveOneGroupOut,
    StratifiedKFold,
    cross_val_predict,
)
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler


SCRIPT_DIR = Path(__file__).resolve().parent
N_FLUX_POINTS = 1024
MATERIAL_SHEETS = ("材料1", "材料2", "材料3", "材料4")
LABEL_TO_CODE = {"正弦波": 1, "三角波": 2, "梯形波": 3}
CODE_TO_LABEL = {value: key for key, value in LABEL_TO_CODE.items()}
PLOT_LABELS = {1: "Sine", 2: "Triangle", 3: "Trapezoid"}
SPECIAL_SAMPLE_IDS = (1, 5, 15, 25, 35, 45, 55, 65, 75, 80)


@dataclass(frozen=True)
class TrainingData:
    waveforms: np.ndarray
    labels: np.ndarray
    materials: np.ndarray


@dataclass(frozen=True)
class TestData:
    sample_ids: np.ndarray
    waveforms: np.ndarray
    materials: np.ndarray


def _numeric_waveforms(frame: pd.DataFrame, source: str) -> np.ndarray:
    """按列位置读取磁通密度，规避四个工作表首列标题不一致的问题。"""
    expected_columns = 4 + N_FLUX_POINTS
    if frame.shape[1] < expected_columns:
        raise ValueError(
            f"{source} 只有 {frame.shape[1]} 列，至少需要 {expected_columns} 列。"
        )

    flux_frame = frame.iloc[:, 4:expected_columns].apply(
        pd.to_numeric, errors="coerce"
    )
    missing_count = int(flux_frame.isna().sum().sum())
    if missing_count:
        raise ValueError(f"{source} 的磁通密度区域含 {missing_count} 个空值或非数值。")

    waveforms = flux_frame.to_numpy(dtype=np.float64, copy=True)
    if waveforms.shape[1] != N_FLUX_POINTS:
        raise ValueError(
            f"{source} 的磁通密度采样点数量为 {waveforms.shape[1]}，"
            f"应为 {N_FLUX_POINTS}。"
        )
    return waveforms


def load_training_data(path: Path) -> TrainingData:
    """读取附件一全部四种材料。"""
    workbook = pd.ExcelFile(path)
    missing_sheets = [name for name in MATERIAL_SHEETS if name not in workbook.sheet_names]
    if missing_sheets:
        raise ValueError(f"附件一缺少工作表：{missing_sheets}")

    waveform_parts: list[np.ndarray] = []
    label_parts: list[np.ndarray] = []
    material_parts: list[np.ndarray] = []

    for sheet_name in MATERIAL_SHEETS:
        frame = pd.read_excel(path, sheet_name=sheet_name)
        waveforms = _numeric_waveforms(frame, f"{path.name}/{sheet_name}")

        raw_labels = frame.iloc[:, 3].astype(str).str.strip()
        unknown_labels = sorted(set(raw_labels) - set(LABEL_TO_CODE))
        if unknown_labels:
            raise ValueError(
                f"{path.name}/{sheet_name} 含未知励磁波形：{unknown_labels}"
            )

        waveform_parts.append(waveforms)
        label_parts.append(raw_labels.map(LABEL_TO_CODE).to_numpy(dtype=np.int64))
        material_parts.append(np.repeat(sheet_name, len(frame)))

    return TrainingData(
        waveforms=np.vstack(waveform_parts),
        labels=np.concatenate(label_parts),
        materials=np.concatenate(material_parts),
    )


def load_test_data(path: Path) -> TestData:
    """读取附件二，保留样本序号和材料，仅按位置读取磁通密度。"""
    frame = pd.read_excel(path)
    waveforms = _numeric_waveforms(frame, path.name)

    sample_ids = pd.to_numeric(frame.iloc[:, 0], errors="coerce")
    if sample_ids.isna().any():
        raise ValueError(f"{path.name} 的样本序号含空值或非数值。")
    sample_ids_array = sample_ids.to_numpy(dtype=np.int64)
    if len(np.unique(sample_ids_array)) != len(sample_ids_array):
        raise ValueError(f"{path.name} 的样本序号不唯一。")

    materials = frame.iloc[:, 3].astype(str).str.strip().to_numpy()
    unknown_materials = sorted(set(materials) - set(MATERIAL_SHEETS))
    if unknown_materials:
        raise ValueError(f"{path.name} 含未知磁芯材料：{unknown_materials}")

    return TestData(
        sample_ids=sample_ids_array,
        waveforms=waveforms,
        materials=materials,
    )


def normalize_waveforms(waveforms: np.ndarray) -> np.ndarray:
    """逐样本去直流并按峰峰值归一化，保留形状、消除幅值量纲。"""
    centered = waveforms - waveforms.mean(axis=1, keepdims=True)
    peak_to_peak = np.ptp(waveforms, axis=1, keepdims=True)
    invalid_rows = np.flatnonzero(peak_to_peak.ravel() <= 1e-12)
    if invalid_rows.size:
        preview = (invalid_rows[:10] + 1).tolist()
        raise ValueError(f"发现近似恒定波形，样本行号示例：{preview}")
    return centered / peak_to_peak


def extract_features(waveforms: np.ndarray) -> tuple[np.ndarray, list[str]]:
    """提取28个分布、斜率、曲率、平台和频谱特征。"""
    normalized = normalize_waveforms(waveforms)
    first_diff = np.diff(normalized, axis=1, append=normalized[:, :1])
    second_diff = np.diff(first_diff, axis=1, append=first_diff[:, :1])
    abs_first_diff = np.abs(first_diff)

    waveform_quantiles = np.quantile(
        normalized, [0.05, 0.25, 0.50, 0.75, 0.95], axis=1
    ).T
    slope_quantiles = np.quantile(
        abs_first_diff, [0.50, 0.90, 0.99], axis=1
    ).T

    spectrum = np.abs(np.fft.rfft(normalized, axis=1))[:, 1:9]
    harmonic_shares = spectrum / (spectrum.sum(axis=1, keepdims=True) + 1e-12)

    plateau_threshold = 0.08 * np.max(
        abs_first_diff, axis=1, keepdims=True
    )

    features = np.column_stack(
        [
            waveform_quantiles,
            normalized.std(axis=1),
            skew(normalized, axis=1, bias=False),
            kurtosis(normalized, axis=1, bias=False),
            np.sqrt(np.mean(normalized**2, axis=1)),
            abs_first_diff.mean(axis=1),
            first_diff.std(axis=1),
            abs_first_diff.max(axis=1),
            slope_quantiles,
            np.abs(second_diff).mean(axis=1),
            second_diff.std(axis=1),
            abs_first_diff.sum(axis=1),
            np.mean(abs_first_diff <= plateau_threshold, axis=1),
            np.mean(np.abs(normalized) >= 0.45, axis=1),
            harmonic_shares,
        ]
    )
    features = np.nan_to_num(features, nan=0.0, posinf=0.0, neginf=0.0)

    names = [
        "B_q05",
        "B_q25",
        "B_q50",
        "B_q75",
        "B_q95",
        "B_std",
        "B_skew",
        "B_kurtosis",
        "B_rms",
        "slope_abs_mean",
        "slope_std",
        "slope_abs_max",
        "slope_abs_q50",
        "slope_abs_q90",
        "slope_abs_q99",
        "curvature_abs_mean",
        "curvature_std",
        "total_variation",
        "plateau_ratio",
        "extreme_ratio",
        *[f"harmonic_{index}_share" for index in range(1, 9)],
    ]
    if features.shape[1] != len(names):
        raise RuntimeError("特征数量和特征名称数量不一致。")
    return features, names


def build_primary_model(seed: int) -> RandomForestClassifier:
    """可解释特征上的主分类器。"""
    return RandomForestClassifier(
        n_estimators=250,
        min_samples_leaf=2,
        class_weight="balanced",
        random_state=seed,
        n_jobs=-1,
    )


def build_consistency_model(seed: int) -> Pipeline:
    """归一化降采样波形上的独立一致性核验模型。"""
    return Pipeline(
        steps=[
            ("scaler", StandardScaler()),
            (
                "classifier",
                LogisticRegression(
                    C=1.0,
                    max_iter=3000,
                    random_state=seed,
                ),
            ),
        ]
    )


def evaluate_model(
    features: np.ndarray,
    labels: np.ndarray,
    materials: np.ndarray,
    seed: int,
) -> tuple[dict, np.ndarray, RandomForestClassifier]:
    """分层五折验证，并执行留一材料验证。"""
    model = build_primary_model(seed)
    cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=seed)
    cv_predictions = cross_val_predict(model, features, labels, cv=cv, n_jobs=1)

    metrics = {
        "sample_count": int(len(labels)),
        "feature_count": int(features.shape[1]),
        "stratified_5fold": {
            "accuracy": float(accuracy_score(labels, cv_predictions)),
            "macro_f1": float(f1_score(labels, cv_predictions, average="macro")),
            "classification_report": classification_report(
                labels,
                cv_predictions,
                labels=[1, 2, 3],
                target_names=[CODE_TO_LABEL[code] for code in (1, 2, 3)],
                output_dict=True,
                zero_division=0,
            ),
        },
        "leave_one_material_out": {},
    }

    logo = LeaveOneGroupOut()
    for train_indices, validation_indices in logo.split(
        features, labels, groups=materials
    ):
        held_out_material = str(np.unique(materials[validation_indices])[0])
        fold_model = clone(model)
        fold_model.fit(features[train_indices], labels[train_indices])
        fold_predictions = fold_model.predict(features[validation_indices])
        metrics["leave_one_material_out"][held_out_material] = {
            "sample_count": int(len(validation_indices)),
            "accuracy": float(
                accuracy_score(labels[validation_indices], fold_predictions)
            ),
            "macro_f1": float(
                f1_score(
                    labels[validation_indices],
                    fold_predictions,
                    average="macro",
                )
            ),
        }

    model.fit(features, labels)
    return metrics, cv_predictions, model


def write_prediction_tables(
    output_dir: Path,
    test_data: TestData,
    predictions: np.ndarray,
    consistency_predictions: np.ndarray,
) -> pd.DataFrame:
    """保存紧凑、可审计的附件二预测表。"""
    result = pd.DataFrame(
        {
            "序号": test_data.sample_ids,
            "磁芯材料": test_data.materials,
            "励磁波形分类结果": predictions.astype(np.int64),
            "一致性模型结果": consistency_predictions.astype(np.int64),
            "两模型是否一致": predictions == consistency_predictions,
        }
    ).sort_values("序号")

    result.to_csv(
        output_dir / "附件二波形分类结果.csv",
        index=False,
        encoding="utf-8-sig",
    )
    result.to_excel(output_dir / "附件二波形分类结果.xlsx", index=False)
    return result


def write_answer_workbook(
    template_path: Path,
    output_path: Path,
    sample_ids: Iterable[int],
    predictions: Iterable[int],
) -> None:
    """复制附件四的结构，在活动工作表第2列按序号填入分类数字。"""
    workbook = load_workbook(template_path)
    worksheet = workbook.active

    row_by_sample_id: dict[int, int] = {}
    last_sample_id: int | None = None
    for row_index in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row_index, column=1).value
        if value is None:
            continue
        try:
            last_sample_id = int(value)
        except (TypeError, ValueError):
            # 官方附件四从A3开始用“=上一行+1”的公式生成序号。
            # openpyxl不会计算公式，因此按连续序号恢复其真实显示值。
            if (
                isinstance(value, str)
                and value.startswith("=")
                and last_sample_id is not None
            ):
                last_sample_id += 1
            else:
                continue
        row_by_sample_id[last_sample_id] = row_index

    # 将模板中的序号公式固化为整数，避免openpyxl保存后公式缓存丢失，
    # 导致Quick Look、网页预览等不计算公式的软件把序号显示为0。
    for sample_id, row_index in row_by_sample_id.items():
        worksheet.cell(row=row_index, column=1, value=int(sample_id))

    missing_ids: list[int] = []
    for sample_id, prediction in zip(sample_ids, predictions):
        sample_id_int = int(sample_id)
        row_index = row_by_sample_id.get(sample_id_int)
        if row_index is None:
            missing_ids.append(sample_id_int)
            continue
        worksheet.cell(row=row_index, column=2, value=int(prediction))

    if missing_ids:
        raise ValueError(f"附件四中找不到以下样本序号：{missing_ids}")

    workbook.save(output_path)

    # 写后校验：确保B2:B81等目标单元格均为真正的整数，而不是文本。
    verified_workbook = load_workbook(output_path, data_only=False, read_only=True)
    verified_sheet = verified_workbook.active
    serial_values = [
        verified_sheet.cell(row=row_index, column=1).value
        for row_index in range(2, verified_sheet.max_row + 1)
    ]
    expected_serial_values = list(range(1, len(serial_values) + 1))
    if serial_values != expected_serial_values:
        verified_workbook.close()
        raise RuntimeError("附件四写后校验失败：序号列不是连续整数。")

    invalid_ids = [
        int(sample_id)
        for sample_id in sample_ids
        if not isinstance(
            verified_sheet.cell(
                row=row_by_sample_id[int(sample_id)],
                column=2,
            ).value,
            int,
        )
    ]
    verified_workbook.close()
    if invalid_ids:
        raise RuntimeError(f"附件四写后校验失败，样本序号：{invalid_ids}")


def plot_confusion(
    labels: np.ndarray,
    predictions: np.ndarray,
    output_path: Path,
) -> None:
    matrix = confusion_matrix(labels, predictions, labels=[1, 2, 3])
    figure, axis = plt.subplots(figsize=(6.4, 5.4))
    image = axis.imshow(matrix, cmap="Blues")
    figure.colorbar(image, ax=axis, fraction=0.046, pad=0.04)

    tick_labels = [PLOT_LABELS[code] for code in (1, 2, 3)]
    axis.set_xticks(range(3), labels=tick_labels)
    axis.set_yticks(range(3), labels=tick_labels)
    axis.set_xlabel("Predicted label")
    axis.set_ylabel("True label")
    axis.set_title("Five-fold cross-validation confusion matrix")

    threshold = matrix.max() / 2
    for row_index in range(3):
        for column_index in range(3):
            value = int(matrix[row_index, column_index])
            axis.text(
                column_index,
                row_index,
                str(value),
                ha="center",
                va="center",
                color="white" if value > threshold else "black",
            )

    figure.tight_layout()
    figure.savefig(output_path, dpi=220)
    plt.close(figure)


def plot_normalized_waveforms(
    waveforms: np.ndarray,
    labels: np.ndarray,
    output_path: Path,
) -> None:
    normalized = normalize_waveforms(waveforms)
    # 将每条波形的最大值循环平移到第0个采样点，避免不同初相位
    # 把中位数曲线“抹平”，使三类形状差异更易解释。
    peak_indices = np.argmax(normalized, axis=1)
    aligned_indices = (
        np.arange(N_FLUX_POINTS)[None, :] + peak_indices[:, None]
    ) % N_FLUX_POINTS
    normalized = np.take_along_axis(normalized, aligned_indices, axis=1)

    x = np.arange(N_FLUX_POINTS)
    colors = {1: "#2F6BFF", 2: "#F59E0B", 3: "#16A085"}

    figure, axes = plt.subplots(1, 3, figsize=(15, 4.5), sharex=True, sharey=True)
    for axis, code in zip(axes, (1, 2, 3)):
        subset = normalized[labels == code]
        median = np.median(subset, axis=0)
        lower = np.quantile(subset, 0.25, axis=0)
        upper = np.quantile(subset, 0.75, axis=0)
        axis.fill_between(x, lower, upper, color=colors[code], alpha=0.20)
        axis.plot(x, median, color=colors[code], linewidth=2)
        axis.set_title(f"{PLOT_LABELS[code]} (n={len(subset)})")
        axis.set_xlabel("Sampling point")
        axis.grid(alpha=0.18)

    axes[0].set_ylabel("Normalized flux density")
    figure.suptitle(
        "Phase-aligned normalized waveform median and interquartile band",
        y=1.02,
    )
    figure.tight_layout()
    figure.savefig(output_path, dpi=220, bbox_inches="tight")
    plt.close(figure)


def plot_feature_importance(
    model: RandomForestClassifier,
    feature_names: list[str],
    output_path: Path,
) -> None:
    importances = model.feature_importances_
    selected = np.argsort(importances)[-12:]

    figure, axis = plt.subplots(figsize=(8, 5.5))
    axis.barh(
        np.arange(len(selected)),
        importances[selected],
        color="#2F6BFF",
        alpha=0.88,
    )
    axis.set_yticks(np.arange(len(selected)), labels=np.array(feature_names)[selected])
    axis.set_xlabel("Random-forest feature importance")
    axis.set_title("Top waveform-shape features")
    axis.grid(axis="x", alpha=0.18)
    figure.tight_layout()
    figure.savefig(output_path, dpi=220)
    plt.close(figure)


def find_default_answer_template() -> Path | None:
    """兼容当前赛题目录，同时允许GitHub用户通过参数显式指定模板。"""
    candidates = (
        SCRIPT_DIR / "附件四（Excel表）_空白模板.xlsx",
    )
    return next((path for path in candidates if path.exists()), None)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--train",
        type=Path,
        default=SCRIPT_DIR / "附件一（训练集）.xlsx",
        help="附件一训练集路径",
    )
    parser.add_argument(
        "--test",
        type=Path,
        default=SCRIPT_DIR / "附件二（测试集）.xlsx",
        help="附件二测试集路径",
    )
    parser.add_argument(
        "--answer-template",
        type=Path,
        default=find_default_answer_template(),
        help="附件四模板路径；找不到时仍会生成独立预测表",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=SCRIPT_DIR,
        help="输出目录",
    )
    parser.add_argument("--seed", type=int, default=42, help="随机种子")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    print("读取并校验附件一、附件二……")
    training_data = load_training_data(args.train.resolve())
    test_data = load_test_data(args.test.resolve())
    print(
        f"训练样本 {len(training_data.labels)} 条，"
        f"测试样本 {len(test_data.sample_ids)} 条。"
    )

    print("提取分布、形状和频域特征……")
    train_features, feature_names = extract_features(training_data.waveforms)
    test_features, test_feature_names = extract_features(test_data.waveforms)
    if feature_names != test_feature_names:
        raise RuntimeError("训练集和测试集的特征定义不一致。")

    print("执行分层五折与留一材料验证……")
    metrics, cv_predictions, primary_model = evaluate_model(
        train_features,
        training_data.labels,
        training_data.materials,
        args.seed,
    )
    primary_predictions = primary_model.predict(test_features).astype(np.int64)

    consistency_model = build_consistency_model(args.seed)
    normalized_train = normalize_waveforms(training_data.waveforms)[:, ::8]
    normalized_test = normalize_waveforms(test_data.waveforms)[:, ::8]
    consistency_model.fit(normalized_train, training_data.labels)
    consistency_predictions = consistency_model.predict(normalized_test).astype(np.int64)

    disagreement_ids = test_data.sample_ids[
        primary_predictions != consistency_predictions
    ].astype(int)
    metrics["test_prediction_consistency"] = {
        "agreement_count": int(
            np.sum(primary_predictions == consistency_predictions)
        ),
        "sample_count": int(len(primary_predictions)),
        "disagreement_sample_ids": disagreement_ids.tolist(),
    }

    prediction_table = write_prediction_tables(
        output_dir,
        test_data,
        primary_predictions,
        consistency_predictions,
    )

    if args.answer_template is not None:
        template_path = args.answer_template.resolve()
        if not template_path.exists():
            raise FileNotFoundError(f"找不到附件四模板：{template_path}")
        answer_output_path = output_dir / "附件四（Excel表）.xlsx"
        write_answer_workbook(
            template_path,
            answer_output_path,
            test_data.sample_ids,
            primary_predictions,
        )
        print(f"附件四已生成：{answer_output_path}")
    else:
        print("未找到附件四模板；已生成独立预测表，可用 --answer-template 指定模板。")

    report_text = classification_report(
        training_data.labels,
        cv_predictions,
        labels=[1, 2, 3],
        target_names=[CODE_TO_LABEL[code] for code in (1, 2, 3)],
        digits=4,
        zero_division=0,
    )
    (output_dir / "五折交叉验证分类报告.txt").write_text(
        report_text, encoding="utf-8"
    )
    (output_dir / "验证指标.json").write_text(
        json.dumps(metrics, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    plot_confusion(
        training_data.labels,
        cv_predictions,
        output_dir / "五折交叉验证混淆矩阵.png",
    )
    plot_normalized_waveforms(
        training_data.waveforms,
        training_data.labels,
        output_dir / "三类归一化波形.png",
    )
    plot_feature_importance(
        primary_model,
        feature_names,
        output_dir / "特征重要性.png",
    )

    counts = (
        prediction_table["励磁波形分类结果"]
        .value_counts()
        .reindex([1, 2, 3], fill_value=0)
    )
    special_results = prediction_table[
        prediction_table["序号"].isin(SPECIAL_SAMPLE_IDS)
    ][["序号", "励磁波形分类结果"]]

    print("\n验证结果")
    print(
        f"五折准确率={metrics['stratified_5fold']['accuracy']:.4f}，"
        f"宏平均F1={metrics['stratified_5fold']['macro_f1']:.4f}"
    )
    for material, material_metrics in metrics["leave_one_material_out"].items():
        print(
            f"留出{material}: 准确率={material_metrics['accuracy']:.4f}，"
            f"宏平均F1={material_metrics['macro_f1']:.4f}"
        )

    print("\n附件二分类数量")
    for code in (1, 2, 3):
        print(f"{code}（{CODE_TO_LABEL[code]}）: {int(counts.loc[code])}")

    print("\n指定样本分类结果")
    print(special_results.to_string(index=False))
    print(
        f"\n两种独立模型一致："
        f"{metrics['test_prediction_consistency']['agreement_count']}/"
        f"{metrics['test_prediction_consistency']['sample_count']}"
    )
    print(f"全部输出位于：{output_dir}")


if __name__ == "__main__":
    main()
